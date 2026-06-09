"""
Stok Toko Elektronik — Inventory Management System v2.0
FastAPI + SQLite + Jinja2 + TailwindCSS
Features: Login, Role, Print, Customer, Debt, Barcode, Backup, Price History, Notifications
"""

import os
import sqlite3
import hashlib
import asyncio
import secrets
import shutil
import json
from datetime import datetime, timedelta
from contextlib import contextmanager
from functools import wraps
from fastapi import FastAPI, Request, Form, Query, Cookie, Response
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Config ──────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "stok.db")
EXPORT_DIR = os.path.join(BASE_DIR, "exports")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
os.makedirs(EXPORT_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

app = FastAPI(title="Stok Toko Elektronik", debug=True)
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

# Add permission check function to Jinja2 globals
def template_has_permission(user, feature):
    """Template helper to check user permissions"""
    if not user:
        return False
    if user["role"] == "bos":
        return True
    with get_db() as db:
        return has_permission(db, user, feature)

templates.env.globals["has_perm"] = template_has_permission

# ── Database ────────────────────────────────────────────────────────────
@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            nama TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('bos', 'og', 'karyawan')),
            aktif INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS kategori (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS produk (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kode TEXT NOT NULL UNIQUE,
            barcode TEXT DEFAULT '',
            nama TEXT NOT NULL,
            kategori_id INTEGER,
            harga_modal REAL NOT NULL DEFAULT 0,
            harga_jual REAL NOT NULL DEFAULT 0,
            stok INTEGER NOT NULL DEFAULT 0,
            stok_minimum INTEGER NOT NULL DEFAULT 5,
            satuan TEXT DEFAULT 'pcs',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (kategori_id) REFERENCES kategori(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS pelanggan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT NOT NULL,
            alamat TEXT DEFAULT '',
            telepon TEXT DEFAULT '',
            email TEXT DEFAULT '',
            catatan TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS stok_mutasi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produk_id INTEGER NOT NULL,
            tipe TEXT NOT NULL CHECK(tipe IN ('masuk', 'keluar')),
            jumlah INTEGER NOT NULL,
            harga_satuan REAL,
            user_id INTEGER,
            keterangan TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (produk_id) REFERENCES produk(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS stok_opname (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produk_id INTEGER NOT NULL,
            stok_sistem INTEGER NOT NULL,
            stok_fisik INTEGER NOT NULL,
            selisih INTEGER NOT NULL,
            user_id INTEGER,
            keterangan TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (produk_id) REFERENCES produk(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS penjualan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pelanggan_id INTEGER,
            user_id INTEGER,
            produk_id INTEGER NOT NULL,
            jumlah INTEGER NOT NULL,
            harga_satuan REAL NOT NULL,
            harga_modal REAL NOT NULL,
            total REAL NOT NULL,
            keuntungan REAL NOT NULL,
            keterangan TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (pelanggan_id) REFERENCES pelanggan(id) ON DELETE SET NULL,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY (produk_id) REFERENCES produk(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS hutang (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pelanggan_id INTEGER,
            penjualan_id INTEGER,
            jumlah REAL NOT NULL DEFAULT 0,
            sudah_bayar REAL NOT NULL DEFAULT 0,
            sisa REAL NOT NULL DEFAULT 0,
            status TEXT DEFAULT 'belum' CHECK(status IN ('belum', 'sebagian', 'lunas')),
            jatuh_tempo DATE,
            keterangan TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (pelanggan_id) REFERENCES pelanggan(id) ON DELETE CASCADE,
            FOREIGN KEY (penjualan_id) REFERENCES penjualan(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS pembayaran_hutang (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hutang_id INTEGER NOT NULL,
            jumlah REAL NOT NULL,
            user_id INTEGER,
            metode_bayar TEXT DEFAULT 'tunai',
            keterangan TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (hutang_id) REFERENCES hutang(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS riwayat_harga (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produk_id INTEGER NOT NULL,
            harga_modal_lama REAL,
            harga_jual_lama REAL,
            harga_modal_baru REAL,
            harga_jual_baru REAL,
            user_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (produk_id) REFERENCES produk(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS notifikasi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipe TEXT NOT NULL,
            pesan TEXT NOT NULL,
            link TEXT DEFAULT '',
            dibaca INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS backups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            ukuran INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS approval (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipe TEXT NOT NULL DEFAULT 'diskon',
            produk_id INTEGER,
            harga_diminta REAL NOT NULL,
            harga_jual REAL NOT NULL,
            harga_bottom REAL NOT NULL,
            karyawan_id INTEGER NOT NULL,
            status TEXT DEFAULT 'pending' CHECK(status IN ('pending', 'approved', 'rejected')),
            approved_by INTEGER,
            alasan TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (produk_id) REFERENCES produk(id) ON DELETE CASCADE,
            FOREIGN KEY (karyawan_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY (approved_by) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT NOT NULL,
            role TEXT NOT NULL,
            aksi TEXT NOT NULL,
            kategori TEXT NOT NULL,
            detail TEXT DEFAULT '',
            ip_address TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS user_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            feature TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(user_id, feature)
        );

        CREATE TABLE IF NOT EXISTS opname_session (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tanggal DATE NOT NULL DEFAULT (DATE('now')),
            status TEXT NOT NULL DEFAULT 'draft',
            catatan TEXT DEFAULT '',
            user_id INTEGER,
            approved_by INTEGER,
            approved_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY (approved_by) REFERENCES users(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS opname_item (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            produk_id INTEGER NOT NULL,
            stok_sistem INTEGER NOT NULL,
            stok_fisik INTEGER NOT NULL DEFAULT 0,
            selisih INTEGER NOT NULL DEFAULT 0,
            keterangan TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES opname_session(id) ON DELETE CASCADE,
            FOREIGN KEY (produk_id) REFERENCES produk(id) ON DELETE CASCADE
        );
        """)

        # ── Migrations for existing DBs
        for col in ['harga_bottom']:
            try: db.execute(f"ALTER TABLE produk ADD COLUMN {col} REAL DEFAULT 0")
            except: pass
        try: db.execute("ALTER TABLE users ADD COLUMN login_at TIMESTAMP")
        except: pass
        try: db.execute("ALTER TABLE users ADD COLUMN session_token TEXT")
        except: pass
        for col in ['nama_customer', 'alamat_customer', 'hp_customer', 'email_customer']:
            try: db.execute(f"ALTER TABLE penjualan ADD COLUMN {col} TEXT DEFAULT ''")
            except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN approval_id INTEGER")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN metode_bayar TEXT DEFAULT 'tunai'")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN status TEXT DEFAULT 'active'")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN void_reason TEXT DEFAULT ''")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN void_by INTEGER")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN void_at TIMESTAMP")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN batch_id TEXT DEFAULT ''")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN diskon REAL DEFAULT 0")
        except: pass
        try: db.execute("ALTER TABLE penjualan ADD COLUMN no_invoice TEXT DEFAULT ''")
        except: pass
        try: db.execute("ALTER TABLE pembayaran_hutang ADD COLUMN metode_bayar TEXT DEFAULT 'tunai'")
        except: pass
        try: db.execute("ALTER TABLE stok_opname ADD COLUMN session_id INTEGER")
        except: pass
        # Init permissions for existing users who don't have any
        try:
            _role_defaults = {
                "bos": {"penjualan": True, "hutang": True, "stok_lihat": True, "stok_masuk": True, "stok_keluar": True, "stok_opname": True, "laporan": True, "audit_log": True, "hapus_data": True},
                "og": {"penjualan": True, "hutang": True, "stok_lihat": True, "stok_masuk": True, "stok_keluar": True, "stok_opname": True, "laporan": True, "audit_log": True, "hapus_data": True},
                "karyawan": {"penjualan": True, "hutang": True, "stok_lihat": True, "stok_masuk": False, "stok_keluar": False, "stok_opname": False, "laporan": False, "lihat_keuntungan": False, "audit_log": False, "hapus_data": False},
            }
            existing_users = db.execute("SELECT id, role FROM users").fetchall()
            for u in existing_users:
                has_perms = db.execute("SELECT COUNT(*) FROM user_permissions WHERE user_id = ?", (u["id"],)).fetchone()[0]
                if has_perms == 0:
                    defaults = _role_defaults.get(u["role"], _role_defaults["karyawan"])
                    for feature, enabled in defaults.items():
                        db.execute("INSERT OR IGNORE INTO user_permissions (user_id, feature, enabled) VALUES (?, ?, ?)", (u["id"], feature, int(enabled)))
        except: pass
        # Make hutang.pelanggan_id nullable (recreate table)
        try:
            schema = db.execute("SELECT sql FROM sqlite_master WHERE name='hutang' AND type='table'").fetchone()
            if schema and "NOT NULL" in schema[0] and "pelanggan_id" in schema[0]:
                db.execute("ALTER TABLE hutang RENAME TO _hutang_migrate_tmp")
                db.execute("""CREATE TABLE hutang (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pelanggan_id INTEGER,
                    penjualan_id INTEGER,
                    jumlah REAL NOT NULL DEFAULT 0,
                    sudah_bayar REAL NOT NULL DEFAULT 0,
                    sisa REAL NOT NULL DEFAULT 0,
                    status TEXT DEFAULT 'belum' CHECK(status IN ('belum', 'sebagian', 'lunas')),
                    jatuh_tempo DATE,
                    keterangan TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (pelanggan_id) REFERENCES pelanggan(id) ON DELETE SET NULL,
                    FOREIGN KEY (penjualan_id) REFERENCES penjualan(id) ON DELETE SET NULL
                )""")
                db.execute("INSERT INTO hutang SELECT * FROM _hutang_migrate_tmp")
                db.execute("DROP TABLE _hutang_migrate_tmp")
        except: pass
        # Migrate role CHECK if needed (add 'og') - safe check first
        has_og = False
        try:
            # Check if 'og' is already in the schema
            schema = db.execute("SELECT sql FROM sqlite_master WHERE name='users' AND type='table'").fetchone()
            if schema and "'og'" in schema[0]:
                has_og = True
        except: pass
        if not has_og:
            try:
                db.execute("ALTER TABLE users RENAME TO _users_migrate_tmp")
                db.execute("""CREATE TABLE users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL, nama TEXT NOT NULL,
                    role TEXT NOT NULL CHECK(role IN ('bos', 'og', 'karyawan')),
                    aktif INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
                db.execute("INSERT INTO users SELECT * FROM _users_migrate_tmp")
                db.execute("DROP TABLE _users_migrate_tmp")
            except:
                # If anything fails, try to restore
                try:
                    db.execute("DROP TABLE IF EXISTS users")
                    db.execute("ALTER TABLE _users_migrate_tmp RENAME TO users")
                except: pass

        # Seed default admin if not exists
        admin = db.execute("SELECT * FROM users WHERE username='admin'").fetchone()
        if not admin:
            pw_hash = hashlib.sha256("admin123".encode()).hexdigest()
            db.execute("INSERT INTO users (username, password_hash, nama, role) VALUES (?, ?, ?, ?)",
                       ("admin", pw_hash, "Administrator", "bos"))

        # Seed default categories
        kategori_list = ["TV & Monitor", "Kulkas & Freezer", "Mesin Cuci", "AC & Kipas",
                         "Kompor & Oven", "Blender & Mixer", "Setrika & Pengering", "Lainnya"]
        for k in kategori_list:
            db.execute("INSERT OR IGNORE INTO kategori (nama) VALUES (?)", (k,))

        # Seed sample data if empty
        if db.execute("SELECT COUNT(*) FROM produk").fetchone()[0] == 0:
            # Sample products
            sample_produk = [
                ("TV LED Samsung 43\"", "TV & Monitor", 3000000, 3500000, 3200000, 15, 5),
                ("Kulkas Samsung 2 Pintu", "Kulkas & Freezer", 3500000, 4200000, 3800000, 8, 3),
                ("Mesin Cuci LG 7kg", "Mesin Cuci", 2500000, 3100000, 2700000, 10, 3),
                ("AC Daikin 1.5PK", "AC & Kipas", 4000000, 4800000, 4300000, 6, 2),
                ("Kompor Gas Rinnai 2 Tungku", "Kompor & Oven", 500000, 650000, 550000, 20, 5),
                ("Blender Philips HR2157", "Blender & Mixer", 350000, 450000, 380000, 25, 5),
                ("Setrika Maspion 300W", "Setrika & Pengering", 150000, 200000, 170000, 30, 10),
                ("TV LED LG 32\"", "TV & Monitor", 2000000, 2500000, 2200000, 12, 5),
                ("Freezer Chest Akari 200L", "Kulkas & Freezer", 1800000, 2300000, 2000000, 5, 2),
                ("Kipas Angin Miyako", "AC & Kipas", 200000, 280000, 230000, 40, 10),
            ]
            for nama, kat, modal, jual, bottom, stok, min_stok in sample_produk:
                kat_id = db.execute("SELECT id FROM kategori WHERE nama=?", (kat,)).fetchone()
                last = db.execute("SELECT id FROM produk ORDER BY id DESC LIMIT 1").fetchone()
                kode = f"P{str((last['id'] if last else 0) + 1).zfill(4)}"
                db.execute("""INSERT INTO produk (kode, barcode, nama, kategori_id, harga_modal, harga_jual, harga_bottom, stok, stok_minimum, satuan)
                    VALUES (?, '', ?, ?, ?, ?, ?, ?, ?, 'pcs')""",
                    (kode, nama, kat_id[0] if kat_id else None, modal, jual, bottom, stok, min_stok))

            # Sample users (OG)
            og_hash = hashlib.sha256("og123".encode()).hexdigest()
            db.execute("INSERT OR IGNORE INTO users (username, password_hash, nama, role) VALUES (?, ?, ?, ?)",
                       ("og1", og_hash, "Tangan Kanan", "og"))
            emp_hash = hashlib.sha256("emp123".encode()).hexdigest()
            db.execute("INSERT OR IGNORE INTO users (username, password_hash, nama, role) VALUES (?, ?, ?, ?)",
                       ("kasir1", emp_hash, "Kasir Satu", "karyawan"))

            # Sample penjualan (last 7 days)
            import random
            today = datetime.now()
            for i in range(15):
                days_ago = random.randint(0, 7)
                date = (today - timedelta(days=days_ago)).strftime("%Y-%m-%d %H:%M:%S")
                prod_id = random.randint(1, 10)
                prod = db.execute("SELECT * FROM produk WHERE id=?", (prod_id,)).fetchone()
                if prod:
                    qty = random.randint(1, 3)
                    harga = prod["harga_jual"]
                    total = harga * qty
                    keuntungan = (harga - prod["harga_modal"]) * qty
                    customer_names = ["Budi Santoso", "Siti Rahayu", "Ahmad Hidayat", "Dewi Lestari", "Rudi Hermawan", "-", "-", "-"]
                    cust = random.choice(customer_names)
                    db.execute("""INSERT INTO penjualan (user_id, produk_id, jumlah, harga_satuan, harga_modal, total, keuntungan, keterangan, nama_customer, created_at)
                        VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (prod_id, qty, harga, prod["harga_modal"], total, keuntungan, "", cust, date))
                    db.execute("UPDATE produk SET stok = stok - ? WHERE id=?", (qty, prod_id))

            # Sample hutang (2 entries)
            hutang_entries = [
                (1, 2500000, 0, 2500000, "belum", (today + timedelta(days=14)).strftime("%Y-%m-%d"), "Budi Santoso"),
                (3, 1500000, 0, 1500000, "belum", (today + timedelta(days=7)).strftime("%Y-%m-%d"), "Ahmad Hidayat"),
            ]
            for pen_id, jumlah, sudah, sisa, status, tempo, cust in hutang_entries:
                db.execute("""INSERT INTO hutang (pelanggan_id, penjualan_id, jumlah, sudah_bayar, sisa, status, jatuh_tempo, keterangan)
                    VALUES (NULL, ?, ?, ?, ?, ?, ?, ?)""", (pen_id, jumlah, sudah, sisa, status, tempo, f"Customer: {cust}"))

            # Sample pelanggan
            sample_pelanggan = [
                ("Budi Santoso", "Jl. Merdeka No. 10, Jakarta", "081234567890", "budi@email.com"),
                ("Siti Rahayu", "Jl. Sudirman No. 25, Bandung", "085678901234", "siti@email.com"),
                ("Ahmad Hidayat", "Jl. Gatot Subroto No. 15, Surabaya", "087890123456", "ahmad@email.com"),
                ("Dewi Lestari", "Jl. Diponegoro No. 8, Semarang", "089012345678", "dewi@email.com"),
                ("Rudi Hermawan", "Jl. Ahmad Yani No. 33, Yogyakarta", "081123456789", "rudi@email.com"),
            ]
            for nama, alamat, telp, email in sample_pelanggan:
                db.execute("INSERT INTO pelanggan (nama, alamat, telepon, email) VALUES (?, ?, ?, ?)",
                           (nama, alamat, telp, email))

            # Sample stok masuk (using stok_mutasi with tipe='masuk')
            sample_masuk = [
                (1, 10, "Restock awal", 2), (2, 5, "Restock awal", 2), (3, 8, "Restock awal", 2),
                (4, 6, "Restock awal", 2), (5, 15, "Restock awal", 2), (6, 20, "Restock awal", 2),
                (7, 25, "Restock awal", 2), (8, 10, "Restock awal", 2), (9, 5, "Restock awal", 2),
                (10, 30, "Restock awal", 2),
            ]
            for prod_id, qty, ket, user_id in sample_masuk:
                days_ago = random.randint(1, 5)
                date = (today - timedelta(days=days_ago)).strftime("%Y-%m-%d %H:%M:%S")
                db.execute("INSERT INTO stok_mutasi (produk_id, tipe, jumlah, keterangan, user_id, created_at) VALUES (?, 'masuk', ?, ?, ?, ?)",
                           (prod_id, qty, ket, user_id, date))

            # Sample stok keluar (using stok_mutasi with tipe='keluar')
            sample_keluar = [
                (1, 2, "Rusak/retur", 2), (6, 3, "Kadaluarsa sample", 2),
                (10, 4, "Contoh display", 2),
            ]
            for prod_id, qty, ket, user_id in sample_keluar:
                days_ago = random.randint(1, 3)
                date = (today - timedelta(days=days_ago)).strftime("%Y-%m-%d %H:%M:%S")
                db.execute("INSERT INTO stok_mutasi (produk_id, tipe, jumlah, keterangan, user_id, created_at) VALUES (?, 'keluar', ?, ?, ?, ?)",
                           (prod_id, qty, ket, user_id, date))

init_db()

# ── Helpers ─────────────────────────────────────────────────────────────
def rupiah(val):
    if val is None: return "Rp 0"
    return f"Rp {val:,.0f}".replace(",", ".")

def format_date(val):
    if val is None: return "-"
    if isinstance(val, str):
        try:
            val = datetime.fromisoformat(val.replace("Z", "+00:00"))
        except (ValueError, TypeError):
            return val
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    return str(val)

def format_date_short(val):
    if val is None: return "-"
    if isinstance(val, str):
        try:
            val = datetime.fromisoformat(val.replace("Z", "+00:00"))
        except (ValueError, TypeError):
            return val
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    return str(val)

templates.env.filters["rupiah"] = rupiah
templates.env.filters["format_date"] = format_date
templates.env.filters["format_date_short"] = format_date_short
templates.env.globals["now"] = datetime.now().strftime("%d/%m/%Y %H:%M")

# ── Auth Helpers ────────────────────────────────────────────────────────
def get_current_user(request: Request):
    token = request.cookies.get("session_token")
    if not token:
        return None
    with get_db() as db:
        user = db.execute("SELECT * FROM users WHERE username = ?", (token,)).fetchone()
    return user

def require_auth(f):
    @wraps(f)
    async def decorated(request: Request, *args, **kwargs):
        user = get_current_user(request)
        if not user:
            return RedirectResponse("/login", status_code=303)
        if user['login_at']:
            login_time = datetime.fromisoformat(user['login_at'])
            if datetime.now() - login_time > timedelta(hours=8):
                with get_db() as db:
                    db.execute("UPDATE users SET session_token=NULL, login_at=NULL WHERE id=?", (user['id'],))
                response = RedirectResponse("/login", status_code=303)
                response.delete_cookie("session_token")
                return response
        request.state.user = user
        if asyncio.iscoroutinefunction(f):
            return await f(request, *args, **kwargs)
        return f(request, *args, **kwargs)
    return decorated

def require_bos(f):
    @wraps(f)
    async def decorated(request: Request, *args, **kwargs):
        user = get_current_user(request)
        if not user:
            return RedirectResponse("/login", status_code=303)
        if user['login_at']:
            login_time = datetime.fromisoformat(user['login_at'])
            if datetime.now() - login_time > timedelta(hours=8):
                with get_db() as db:
                    db.execute("UPDATE users SET session_token=NULL, login_at=NULL WHERE id=?", (user['id'],))
                response = RedirectResponse("/login", status_code=303)
                response.delete_cookie("session_token")
                return response
        if user["role"] != "bos":
            return RedirectResponse("/?error=unauthorized", status_code=303)
        request.state.user = user
        if asyncio.iscoroutinefunction(f):
            return await f(request, *args, **kwargs)
        return f(request, *args, **kwargs)
    return decorated

def require_bos_or_og(f):
    @wraps(f)
    async def decorated(request: Request, *args, **kwargs):
        user = get_current_user(request)
        if not user:
            return RedirectResponse("/login", status_code=303)
        if user['login_at']:
            login_time = datetime.fromisoformat(user['login_at'])
            if datetime.now() - login_time > timedelta(hours=8):
                with get_db() as db:
                    db.execute("UPDATE users SET session_token=NULL, login_at=NULL WHERE id=?", (user['id'],))
                response = RedirectResponse("/login", status_code=303)
                response.delete_cookie("session_token")
                return response
        if user["role"] not in ("bos", "og"):
            return RedirectResponse("/?error=unauthorized", status_code=303)
        request.state.user = user
        if asyncio.iscoroutinefunction(f):
            return await f(request, *args, **kwargs)
        return f(request, *args, **kwargs)
    return decorated

# ── Permission Helper ──────────────────────────────────────────────────
ALL_FEATURES = [
    "penjualan", "hutang",
    "stok_lihat", "stok_masuk", "stok_keluar", "stok_opname",
    "laporan", "lihat_keuntungan", "audit_log", "hapus_data",
]

ROLE_DEFAULTS = {
    "bos": {f: True for f in ALL_FEATURES},
    "og": {f: True for f in ALL_FEATURES},
    "karyawan": {
            "penjualan": True, "hutang": True,
            "stok_lihat": True, "stok_masuk": False, "stok_keluar": False, "stok_opname": False,
            "laporan": False, "lihat_keuntungan": False, "audit_log": False, "hapus_data": False,
        },
}

def get_user_permissions(db, user_id):
    perms = db.execute("SELECT feature, enabled FROM user_permissions WHERE user_id = ?", (user_id,)).fetchall()
    if perms:
        return {p["feature"]: bool(p["enabled"]) for p in perms}
    return {}

def has_permission(db, user, feature):
    if not user:
        return False
    if user["role"] == "bos":
        return True
    perms = get_user_permissions(db, user["id"])
    if not perms:
        defaults = ROLE_DEFAULTS.get(user["role"], {})
        return defaults.get(feature, False)
    return perms.get(feature, False)

def init_user_permissions(db, user_id, role):
    defaults = ROLE_DEFAULTS.get(role, {})
    for feature, enabled in defaults.items():
        db.execute(
            "INSERT OR IGNORE INTO user_permissions (user_id, feature, enabled) VALUES (?, ?, ?)",
            (user_id, feature, int(enabled))
        )

# ── Notification Helper ─────────────────────────────────────────────────
def log_audit(db, user, aksi, kategori, detail="", ip=""):
    """Log user activity for audit trail"""
    if user:
        db.execute(
            "INSERT INTO audit_log (user_id, username, role, aksi, kategori, detail, ip_address) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user['id'], user['username'], user['role'], aksi, kategori, detail, ip)
        )

def generate_no_invoice(db):
    """Generate invoice number: YY/MON/INV-XXXX (e.g. 26/JUN/INV-0001)"""
    now = datetime.now()
    year_short = now.strftime("%y")  # 2-digit year
    month_upper = now.strftime("%b").upper()  # JAN, FEB, MAR, etc.
    prefix = f"{year_short}/{month_upper}/INV-"

    # Count existing invoices this month
    row = db.execute(
        "SELECT COUNT(*) FROM penjualan WHERE no_invoice LIKE ?",
        (f"{prefix}%",)
    ).fetchone()
    seq = (row[0] or 0) + 1

    return f"{prefix}{seq:04d}"

def add_notif(tipe, pesan, link=""):
    with get_db() as db:
        db.execute("INSERT INTO notifikasi (tipe, pesan, link) VALUES (?, ?, ?)", (tipe, pesan, link))

def check_low_stock():
    with get_db() as db:
        low = db.execute("SELECT * FROM produk WHERE stok <= stok_minimum AND stok > 0").fetchall()
        out = db.execute("SELECT * FROM produk WHERE stok = 0").fetchall()
        for p in low:
            existing = db.execute("SELECT * FROM notifikasi WHERE pesan LIKE ? AND dibaca=0",
                                  (f"%{p['nama']}%",)).fetchone()
            if not existing:
                add_notif("stok_menipis", f"Stok {p['nama']} tinggal {p['stok']} {p['satuan']}", "/produk")
        for p in out:
            existing = db.execute("SELECT * FROM notifikasi WHERE pesan LIKE ? AND dibaca=0",
                                  (f"%{p['nama']}%habis%",)).fetchone()
            if not existing:
                add_notif("stok_habis", f"Stok {p['nama']} HABIS!", "/produk")

# ── Backup Helper ───────────────────────────────────────────────────────
def auto_backup():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"stok_backup_{timestamp}.db"
    filepath = os.path.join(BACKUP_DIR, filename)
    shutil.copy2(DB_PATH, filepath)
    ukuran = os.path.getsize(filepath)
    with get_db() as db:
        db.execute("INSERT INTO backups (filename, ukuran) VALUES (?, ?)", (filename, ukuran))
    # Keep only last 10 backups
    with get_db() as db:
        old = db.execute("SELECT * FROM backups ORDER BY created_at DESC LIMIT 10 OFFSET 10").fetchall()
        for b in old:
            old_path = os.path.join(BACKUP_DIR, b["filename"])
            if os.path.exists(old_path):
                os.remove(old_path)
            db.execute("DELETE FROM backups WHERE id=?", (b["id"],))
    return filepath

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: AUTH
# ═══════════════════════════════════════════════════════════════════════
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, error: str = ""):
    return templates.TemplateResponse(request, "login.html", {"error": error})

@app.post("/login")
def login(request: Request, response: Response, username: str = Form(...), password: str = Form(...)):
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    with get_db() as db:
        user = db.execute("SELECT * FROM users WHERE username=? AND password_hash=? AND aktif=1",
                          (username, pw_hash)).fetchone()
        if not user:
            return templates.TemplateResponse(request, "login.html", {"error": "Username atau password salah!"})
        token = user["username"]
        db.execute("UPDATE users SET session_token=?, login_at=? WHERE id=?", (token, datetime.now().isoformat(), user['id']))
        log_audit(db, user, "Login", "autentikasi", "Login berhasil", request.client.host)
    response = RedirectResponse("/", status_code=303)
    response.set_cookie("session_token", token, max_age=86400)
    return response

@app.get("/logout")
def logout(request: Request, reason: str = ""):
    user = get_current_user(request)
    if user:
        with get_db() as db:
            if reason == "timeout":
                log_audit(db, user, "Auto Logout", "autentikasi", "Session expired (8 jam)", request.client.host)
            else:
                log_audit(db, user, "Logout", "autentikasi", "Manual logout", request.client.host)
            db.execute("UPDATE users SET session_token=NULL, login_at=NULL WHERE id=?", (user['id'],))
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie("session_token")
    if reason == "timeout":
        response.set_cookie("flash_msg", "Session expired. Silakan login kembali.", max_age=5)
    return response

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════
@app.get("/", response_class=HTMLResponse)
@require_auth
def dashboard(request: Request):
    user = request.state.user
    check_low_stock()

    with get_db() as db:
        total_produk = db.execute("SELECT COUNT(*) FROM produk").fetchone()[0]
        total_stok = db.execute("SELECT COALESCE(SUM(stok), 0) FROM produk").fetchone()[0]
        stok_menipis = db.execute("SELECT COUNT(*) FROM produk WHERE stok <= stok_minimum").fetchone()[0]

        today = datetime.now().strftime("%Y-%m-%d")
        penjualan_hari_ini = db.execute(
            "SELECT COALESCE(SUM(total), 0), COALESCE(SUM(keuntungan), 0), COUNT(*) FROM penjualan WHERE DATE(created_at) = ?",
            (today,)).fetchone()

        # Breakdown metode bayar hari ini
        bayar_hari_ini = {}
        for metode in ['tunai', 'transfer', 'hutang']:
            row = db.execute(
                "SELECT COALESCE(SUM(total), 0), COUNT(*) FROM penjualan WHERE DATE(created_at) = ? AND metode_bayar = ?",
                (today, metode)).fetchone()
            bayar_hari_ini[metode] = {'total': row[0], 'count': row[1]}

        bulan_ini = datetime.now().strftime("%Y-%m")
        penjualan_bulan = db.execute(
            "SELECT COALESCE(SUM(total), 0), COALESCE(SUM(keuntungan), 0), COUNT(*) FROM penjualan WHERE strftime('%Y-%m', created_at) = ?",
            (bulan_ini,)).fetchone()

        minggu_lalu = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        penjualan_minggu = db.execute(
            "SELECT COALESCE(SUM(total), 0), COALESCE(SUM(keuntungan), 0), COUNT(*) FROM penjualan WHERE DATE(created_at) >= ?",
            (minggu_lalu,)).fetchone()

        # Breakdown metode bayar minggu ini
        bayar_minggu_ini = {}
        for metode in ['tunai', 'transfer', 'hutang']:
            row = db.execute(
                "SELECT COALESCE(SUM(total), 0), COUNT(*) FROM penjualan WHERE DATE(created_at) >= ? AND metode_bayar = ?",
                (minggu_lalu, metode)).fetchone()
            bayar_minggu_ini[metode] = {'total': row[0], 'count': row[1]}

        produk_menipis = db.execute("""
            SELECT p.*, k.nama as kategori_nama FROM produk p
            LEFT JOIN kategori k ON p.kategori_id = k.id
            WHERE p.stok <= p.stok_minimum ORDER BY p.stok ASC LIMIT 10
        """).fetchall()

        mutasi_terbaru = db.execute("""
            SELECT m.*, p.nama as produk_nama, p.kode as produk_kode
            FROM stok_mutasi m JOIN produk p ON m.produk_id = p.id
            ORDER BY m.created_at DESC LIMIT 10
        """).fetchall()

        chart_labels, chart_sales, chart_profit = [], [], []
        for i in range(6, -1, -1):
            d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            chart_labels.append((datetime.now() - timedelta(days=i)).strftime("%d/%m"))
            row = db.execute(
                "SELECT COALESCE(SUM(total), 0), COALESCE(SUM(keuntungan), 0) FROM penjualan WHERE DATE(created_at) = ?",
                (d,)).fetchone()
            chart_sales.append(row[0])
            chart_profit.append(row[1])

        top_produk = db.execute("""
            SELECT p.nama, SUM(pen.jumlah) as total_qty, SUM(pen.total) as total_rp
            FROM penjualan pen JOIN produk p ON pen.produk_id = p.id
            WHERE strftime('%Y-%m', pen.created_at) = ?
            GROUP BY p.id ORDER BY total_qty DESC LIMIT 5
        """, (bulan_ini,)).fetchall()

        # Notifications
        notif_count = db.execute("SELECT COUNT(*) FROM notifikasi WHERE dibaca=0").fetchone()[0]
        notif_list = db.execute("SELECT * FROM notifikasi ORDER BY created_at DESC LIMIT 10").fetchall()

        # Hutang jatuh tempo
        hutang_jatuh = db.execute("""
            SELECT h.*, p.nama as pelanggan_nama FROM hutang h
            JOIN pelanggan p ON h.pelanggan_id = p.id
            WHERE h.status != 'lunas' AND h.jatuh_tempo <= date('now', '+7 days')
            ORDER BY h.jatuh_tempo ASC LIMIT 5
        """).fetchall()

    return templates.TemplateResponse(request, "dashboard.html", {
        "user": user, "total_produk": total_produk, "total_stok": total_stok,
        "stok_menipis": stok_menipis, "penjualan_hari_ini": penjualan_hari_ini,
        "penjualan_bulan": penjualan_bulan, "penjualan_minggu": penjualan_minggu,
        "produk_menipis": produk_menipis, "mutasi_terbaru": mutasi_terbaru,
        "chart_labels": chart_labels, "chart_sales": chart_sales, "chart_profit": chart_profit,
        "top_produk": top_produk, "notif_count": notif_count, "notif_list": notif_list,
        "hutang_jatuh": hutang_jatuh,
        "bayar_hari_ini": bayar_hari_ini, "bayar_minggu_ini": bayar_minggu_ini,
    })

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: NOTIFIKASI
# ═══════════════════════════════════════════════════════════════════════
@app.get("/notifikasi", response_class=HTMLResponse)
@require_auth
def notifikasi_page(request: Request):
    with get_db() as db:
        notif = db.execute("SELECT * FROM notifikasi ORDER BY created_at DESC LIMIT 50").fetchall()
    return templates.TemplateResponse(request, "notifikasi.html", {"request": request, "user": request.state.user, "notif": notif})

@app.post("/notifikasi/baca")
@require_auth
def notifikasi_baca(request: Request):
    with get_db() as db:
        db.execute("UPDATE notifikasi SET dibaca=1")
    return RedirectResponse("/notifikasi", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: PRODUK
# ═══════════════════════════════════════════════════════════════════════
@app.get("/produk", response_class=HTMLResponse)
@require_auth
def produk_list(request: Request, q: str = "", kategori: int = 0, sort: str = "az"):
    with get_db() as db:
        query = "SELECT p.*, k.nama as kategori_nama FROM produk p LEFT JOIN kategori k ON p.kategori_id = k.id WHERE 1=1"
        params = []
        if q:
            query += " AND (p.nama LIKE ? OR p.barcode LIKE ?)"
            params.extend([f"%{q}%", f"%{q}%"])
        if kategori:
            query += " AND p.kategori_id = ?"
            params.append(kategori)
        sort_map = {
            "az": "p.nama ASC",
            "za": "p.nama DESC",
            "harga_asc": "p.harga_jual ASC",
            "harga_desc": "p.harga_jual DESC",
            "stok_asc": "p.stok ASC",
            "stok_desc": "p.stok DESC",
        }
        order = sort_map.get(sort, "p.nama ASC")
        query += f" ORDER BY {order}"
        produk = db.execute(query, params).fetchall()
        kategori_list = db.execute("SELECT * FROM kategori ORDER BY nama").fetchall()
    return templates.TemplateResponse(request, "produk.html", {
        "request": request, "user": request.state.user, "produk": produk,
        "kategori_list": kategori_list, "q": q, "kategori": kategori, "sort": sort
    })

@app.get("/api/produk/barcode/{barcode}")
@require_auth
def api_produk_barcode(request: Request, barcode: str):
    with get_db() as db:
        produk = db.execute("SELECT * FROM produk WHERE barcode = ?", (barcode,)).fetchone()
    if produk:
        return JSONResponse({"found": True, "id": produk["id"], "kode": produk["kode"],
                             "nama": produk["nama"], "harga_jual": produk["harga_jual"],
                             "harga_bottom": produk["harga_bottom"] or 0, "stok": produk["stok"]})
    return JSONResponse({"found": False})

@app.get("/api/produk/{id}")
@require_auth
def api_produk_detail(request: Request, id: int):
    with get_db() as db:
        produk = db.execute("SELECT id, nama, kode, harga_jual, harga_bottom, harga_modal, stok FROM produk WHERE id = ?", (id,)).fetchone()
    if not produk:
        return JSONResponse({"found": False})
    return JSONResponse({
        "found": True, "id": produk["id"], "nama": produk["nama"],
        "harga_jual": produk["harga_jual"], "harga_bottom": produk["harga_bottom"] or 0,
        "harga_modal": produk["harga_modal"], "stok": produk["stok"]
    })

@app.post("/produk/tambah")
@require_auth
def produk_tambah(request: Request, barcode: str = Form(""),
                  nama: str = Form(...), kategori_id: int = Form(0),
                  harga_modal: float = Form(0), harga_jual: float = Form(0),
                  harga_bottom: float = Form(0),
                  stok: int = Form(0), stok_minimum: int = Form(5), satuan: str = Form("pcs")):
    user = request.state.user
    with get_db() as db:
        if not has_permission(db, user, "hapus_data"):
            return RedirectResponse("/produk", status_code=303)
        last = db.execute("SELECT id FROM produk ORDER BY id DESC LIMIT 1").fetchone()
        kode = f"P{str((last['id'] if last else 0) + 1).zfill(4)}"
        db.execute("""
            INSERT INTO produk (kode, barcode, nama, kategori_id, harga_modal, harga_jual, harga_bottom, stok, stok_minimum, satuan)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (kode, barcode, nama, kategori_id if kategori_id else None, harga_modal, harga_jual, harga_bottom, stok, stok_minimum, satuan))
        log_audit(db, request.state.user, "Tambah Produk", "produk", f"{nama} - Harga: Rp {harga_jual:,.0f}", request.client.host)
    return RedirectResponse("/produk", status_code=303)

@app.post("/produk/edit/{id}")
@require_auth
def produk_edit(request: Request, id: int, barcode: str = Form(""),
                nama: str = Form(...), kategori_id: int = Form(0),
                harga_modal: float = Form(0), harga_jual: float = Form(0),
                harga_bottom: float = Form(0),
                stok_minimum: int = Form(5), satuan: str = Form("pcs")):
    user = request.state.user
    with get_db() as db:
        if not has_permission(db, user, "hapus_data"):
            return RedirectResponse("/produk", status_code=303)
        old = db.execute("SELECT * FROM produk WHERE id=?", (id,)).fetchone()
        if old and (old["harga_modal"] != harga_modal or old["harga_jual"] != harga_jual):
            db.execute("""
                INSERT INTO riwayat_harga (produk_id, harga_modal_lama, harga_jual_lama, harga_modal_baru, harga_jual_baru, user_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (id, old["harga_modal"], old["harga_jual"], harga_modal, harga_jual, request.state.user["id"]))
        db.execute("""
            UPDATE produk SET barcode=?, nama=?, kategori_id=?, harga_modal=?, harga_jual=?,
            harga_bottom=?, stok_minimum=?, satuan=?, updated_at=CURRENT_TIMESTAMP WHERE id=?
        """, (barcode, nama, kategori_id if kategori_id else None, harga_modal, harga_jual, harga_bottom, stok_minimum, satuan, id))
        log_audit(db, request.state.user, "Edit Produk", "produk", f"ID {id}: {nama}", request.client.host)
    return RedirectResponse("/produk", status_code=303)

@app.get("/produk/hapus/{id}")
@require_bos_or_og
def produk_hapus(request: Request, id: int):
    user = request.state.user
    with get_db() as db:
        if not has_permission(db, user, "hapus_data"):
            return RedirectResponse("/produk", status_code=303)
        old = db.execute("SELECT * FROM produk WHERE id=?", (id,)).fetchone()
        if old:
            log_audit(db, request.state.user, "Hapus Produk", "produk", f"ID {id}: {old['nama']}", request.client.host)
        db.execute("DELETE FROM produk WHERE id=?", (id,))
    return RedirectResponse("/produk", status_code=303)

@app.get("/produk/{id}/riwayat-harga", response_class=HTMLResponse)
@require_auth
def riwayat_harga_page(request: Request, id: int):
    with get_db() as db:
        produk = db.execute("SELECT * FROM produk WHERE id=?", (id,)).fetchone()
        riwayat = db.execute("""
            SELECT r.*, u.nama as user_nama FROM riwayat_harga r
            LEFT JOIN users u ON r.user_id = u.id
            WHERE r.produk_id=? ORDER BY r.created_at DESC
        """, (id,)).fetchall()
    return templates.TemplateResponse(request, "riwayat_harga.html", {
        "request": request, "user": request.state.user, "produk": produk, "riwayat": riwayat
    })

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: STOK (Overview Semua Barang)
# ═══════════════════════════════════════════════════════════════════════
@app.get("/stok", response_class=HTMLResponse)
@require_auth
def stok_page(request: Request, q: str = "", filter_kategori: str = "", filter_kode: str = "", tgl_dari: str = "", tgl_sampai: str = ""):
    with get_db() as db:
        kategori_list = db.execute("SELECT * FROM kategori ORDER BY nama").fetchall()

        # Build date-filtered subqueries
        masuk_where = "tipe = 'masuk'"
        keluar_where = "tipe = 'keluar'"
        masuk_params = []
        keluar_params = []
        if tgl_dari:
            masuk_where += " AND DATE(created_at) >= ?"
            keluar_where += " AND DATE(created_at) >= ?"
            masuk_params.append(tgl_dari)
            keluar_params.append(tgl_dari)
        if tgl_sampai:
            masuk_where += " AND DATE(created_at) <= ?"
            keluar_where += " AND DATE(created_at) <= ?"
            masuk_params.append(tgl_sampai)
            keluar_params.append(tgl_sampai)

        masuk_sub = f"(SELECT produk_id as pid, SUM(jumlah) as total_masuk FROM stok_mutasi WHERE {masuk_where} GROUP BY produk_id)"
        keluar_sub = f"(SELECT produk_id as pid, SUM(jumlah) as total_keluar FROM stok_mutasi WHERE {keluar_where} GROUP BY produk_id)"

        query = f"""
            SELECT p.id, p.kode, k.nama as kategori_nama, p.nama,
                   COALESCE(masuk.total_masuk, 0) as terima,
                   COALESCE(keluar.total_keluar, 0) as keluar,
                   p.stok as stok_akhir,
                   (SELECT MAX(DATE(sm.created_at)) FROM stok_mutasi sm WHERE sm.produk_id = p.id) as tgl_transaksi
            FROM produk p
            LEFT JOIN kategori k ON p.kategori_id = k.id
            LEFT JOIN {masuk_sub} masuk ON masuk.pid = p.id
            LEFT JOIN {keluar_sub} keluar ON keluar.pid = p.id
            WHERE 1=1
        """
        # All params: masuk subquery params + keluar subquery params + main filters
        all_params = masuk_params + keluar_params

        if q:
            query += " AND p.nama LIKE ?"
            all_params.append(f"%{q}%")
        if filter_kategori:
            query += " AND k.nama = ?"
            all_params.append(filter_kategori)
        if filter_kode:
            query += " AND p.kode = ?"
            all_params.append(filter_kode)

        query += " ORDER BY p.nama"
        stok_data = db.execute(query, all_params).fetchall()
        kode_list = db.execute("SELECT DISTINCT kode FROM produk ORDER BY kode").fetchall()

    return templates.TemplateResponse(request, "stok.html", {
        "request": request, "user": request.state.user,
        "stok_data": stok_data, "kategori_list": kategori_list,
        "kode_list": kode_list, "q": q,
        "filter_kategori": filter_kategori, "filter_kode": filter_kode,
        "tgl_dari": tgl_dari, "tgl_sampai": tgl_sampai,
    })

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: STOK MASUK
# ═══════════════════════════════════════════════════════════════════════
@app.get("/stok/masuk", response_class=HTMLResponse)
@require_auth
def stok_masuk_page(request: Request, tgl_dari: str = "", tgl_sampai: str = ""):
    with get_db() as db:
        produk = db.execute("SELECT * FROM produk ORDER BY nama").fetchall()
        where = "m.tipe = 'masuk'"
        params = []
        if tgl_dari:
            where += " AND DATE(m.created_at) >= ?"
            params.append(tgl_dari)
        if tgl_sampai:
            where += " AND DATE(m.created_at) <= ?"
            params.append(tgl_sampai)
        mutasi = db.execute(f"""
            SELECT m.*, p.nama as produk_nama, p.kode as produk_kode, u.nama as user_nama
            FROM stok_mutasi m JOIN produk p ON m.produk_id = p.id
            LEFT JOIN users u ON m.user_id = u.id
            WHERE {where} ORDER BY m.created_at DESC LIMIT 50
        """, params).fetchall()
    return templates.TemplateResponse(request, "stok_masuk.html", {
        "request": request, "user": request.state.user, "produk": produk, "mutasi": mutasi,
        "tgl_dari": tgl_dari, "tgl_sampai": tgl_sampai,
    })

@app.post("/stok/masuk")
@require_auth
def stok_masuk(request: Request, produk_id: int = Form(...), jumlah: int = Form(...),
               harga_satuan: float = Form(0), keterangan: str = Form("")):
    with get_db() as db:
        db.execute("UPDATE produk SET stok = stok + ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (jumlah, produk_id))
        db.execute("""
            INSERT INTO stok_mutasi (produk_id, tipe, jumlah, harga_satuan, user_id, keterangan)
            VALUES (?, 'masuk', ?, ?, ?, ?)
        """, (produk_id, jumlah, harga_satuan, request.state.user["id"], keterangan))
        if harga_satuan > 0:
            old = db.execute("SELECT * FROM produk WHERE id=?", (produk_id,)).fetchone()
            if old and old["harga_modal"] != harga_satuan:
                db.execute("""
                    INSERT INTO riwayat_harga (produk_id, harga_modal_lama, harga_jual_lama, harga_modal_baru, harga_jual_baru, user_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (produk_id, old["harga_modal"], old["harga_jual"], harga_satuan, old["harga_jual"], request.state.user["id"]))
            db.execute("UPDATE produk SET harga_modal = ? WHERE id = ?", (harga_satuan, produk_id))
        log_audit(db, request.state.user, "Stok Masuk", "stok", f"Produk ID {produk_id} +{jumlah}", request.client.host)
    check_low_stock()
    return RedirectResponse("/stok/masuk", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: STOK KELUAR
# ═══════════════════════════════════════════════════════════════════════
@app.get("/stok/keluar", response_class=HTMLResponse)
@require_auth
def stok_keluar_page(request: Request, tgl_dari: str = "", tgl_sampai: str = ""):
    with get_db() as db:
        produk = db.execute("SELECT * FROM produk WHERE stok > 0 ORDER BY nama").fetchall()
        where = "m.tipe = 'keluar'"
        params = []
        if tgl_dari:
            where += " AND DATE(m.created_at) >= ?"
            params.append(tgl_dari)
        if tgl_sampai:
            where += " AND DATE(m.created_at) <= ?"
            params.append(tgl_sampai)
        mutasi = db.execute(f"""
            SELECT m.*, p.nama as produk_nama, p.kode as produk_kode, u.nama as user_nama
            FROM stok_mutasi m JOIN produk p ON m.produk_id = p.id
            LEFT JOIN users u ON m.user_id = u.id
            WHERE {where} ORDER BY m.created_at DESC LIMIT 50
        """, params).fetchall()
    return templates.TemplateResponse(request, "stok_keluar.html", {
        "request": request, "user": request.state.user, "produk": produk, "mutasi": mutasi,
        "tgl_dari": tgl_dari, "tgl_sampai": tgl_sampai,
    })

@app.post("/stok/keluar")
@require_auth
def stok_keluar(request: Request, produk_id: int = Form(...), jumlah: int = Form(...),
                keterangan: str = Form("")):
    with get_db() as db:
        produk = db.execute("SELECT * FROM produk WHERE id = ?", (produk_id,)).fetchone()
        if produk["stok"] < jumlah:
            return RedirectResponse("/stok/keluar?error=stok_kurang", status_code=303)
        db.execute("UPDATE produk SET stok = stok - ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (jumlah, produk_id))
        db.execute("""
            INSERT INTO stok_mutasi (produk_id, tipe, jumlah, harga_satuan, user_id, keterangan)
            VALUES (?, 'keluar', ?, ?, ?, ?)
        """, (produk_id, jumlah, produk["harga_jual"], request.state.user["id"], keterangan))
        log_audit(db, request.state.user, "Stok Keluar", "stok", f"Produk ID {produk_id} -{jumlah} ({keterangan})", request.client.host)
    check_low_stock()
    return RedirectResponse("/stok/keluar", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: APPROVAL
# ═══════════════════════════════════════════════════════════════════════
@app.post("/approval/request")
@require_auth
def approval_request_form(request: Request, produk_id: int = Form(...),
                          harga_diminta: float = Form(...), alasan: str = Form("")):
    user = request.state.user
    with get_db() as db:
        produk = db.execute("SELECT * FROM produk WHERE id=?", (produk_id,)).fetchone()
        if not produk:
            return RedirectResponse("/penjualan?error=produk_not_found", status_code=303)
        db.execute("""
            INSERT INTO approval (tipe, produk_id, harga_diminta, harga_jual, harga_bottom, karyawan_id, alasan)
            VALUES ('diskon', ?, ?, ?, ?, ?, ?)
        """, (produk_id, harga_diminta, produk["harga_jual"], produk["harga_bottom"] or 0, user["id"], alasan))
        # Notify all bos and OG
        db.execute("""
            INSERT INTO notifikasi (tipe, pesan, link) VALUES ('approval', ?, '/approval')
        """, (f"Persetujuan diskon: {produk['nama']} - harga Rp {harga_diminta:,.0f} (oleh {user['nama']})",))
    return RedirectResponse("/penjualan?approval=sent", status_code=303)

@app.get("/approval", response_class=HTMLResponse)
@require_bos_or_og
def approval_page(request: Request):
    with get_db() as db:
        approvals = db.execute("""
            SELECT a.*, p.nama as produk_nama, p.harga_jual, p.harga_bottom, p.harga_modal,
                   u.nama as karyawan_nama, au.nama as approver_nama
            FROM approval a
            JOIN produk p ON a.produk_id = p.id
            JOIN users u ON a.karyawan_id = u.id
            LEFT JOIN users au ON a.approved_by = au.id
            ORDER BY a.created_at DESC LIMIT 50
        """).fetchall()
    return templates.TemplateResponse(request, "approval.html", {
        "request": request, "user": request.state.user, "approvals": approvals
    })

@app.post("/approval/approve/{id}")
@require_bos_or_og
def approval_approve(request: Request, id: int):
    user = request.state.user
    with get_db() as db:
        approval = db.execute("SELECT * FROM approval WHERE id=? AND status='pending'", (id,)).fetchone()
        if approval:
            db.execute("UPDATE approval SET status='approved', approved_by=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                       (user["id"], id))
            db.execute("INSERT INTO notifikasi (tipe, pesan, link) VALUES ('approval_done', ?, '/penjualan')",
                       (f"Diskon DISETUJUI oleh {user['nama']}: Rp {approval['harga_diminta']:,.0f}",))
            log_audit(db, user, "Approve Diskon", "approval", f"Approval #{id}", request.client.host)
    return RedirectResponse("/approval", status_code=303)

@app.post("/approval/reject/{id}")
@require_bos_or_og
def approval_reject(request: Request, id: int):
    user = request.state.user
    with get_db() as db:
        approval = db.execute("SELECT * FROM approval WHERE id=? AND status='pending'", (id,)).fetchone()
        if approval:
            db.execute("UPDATE approval SET status='rejected', approved_by=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                       (user["id"], id))
            db.execute("INSERT INTO notifikasi (tipe, pesan, link) VALUES ('approval_done', ?, '/penjualan')",
                       (f"Diskon DITOLAK oleh {user['nama']}: Rp {approval['harga_diminta']:,.0f}",))
            log_audit(db, user, "Reject Diskon", "approval", f"Approval #{id}", request.client.host)
    return RedirectResponse("/approval", status_code=303)

@app.get("/api/approvals/pending")
@require_auth
def api_pending_approvals(request: Request):
    with get_db() as db:
        count = db.execute("SELECT COUNT(*) FROM approval WHERE status='pending'").fetchone()[0]
    return JSONResponse({"count": count})

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: PENJUALAN
# ═══════════════════════════════════════════════════════════════════════
@app.get("/penjualan", response_class=HTMLResponse)
@require_auth
def penjualan_page(request: Request, periode: str = "hari", tgl_dari: str = "", tgl_sampai: str = ""):
    with get_db() as db:
        today = datetime.now().strftime("%Y-%m-%d")
        bulan_ini = datetime.now().strftime("%Y-%m")
        minggu_lalu = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")

        params = []
        if tgl_dari or tgl_sampai:
            # Custom date range takes priority over periode
            periode = "custom"
            where = "1=1"
            if tgl_dari:
                where += " AND DATE(pen.created_at) >= ?"
                params.append(tgl_dari)
            if tgl_sampai:
                where += " AND DATE(pen.created_at) <= ?"
                params.append(tgl_sampai)
        elif periode == "hari": where = f"DATE(pen.created_at) = '{today}'"
        elif periode == "minggu": where = f"DATE(pen.created_at) >= '{minggu_lalu}'"
        elif periode == "bulan": where = f"strftime('%Y-%m', pen.created_at) = '{bulan_ini}'"
        else: where = "1=1"

        penjualan = db.execute(f"""
            SELECT pen.*, pr.nama as produk_nama, pr.kode as produk_kode,
                   u.nama as user_nama,
                   vu.nama as void_by_nama
            FROM penjualan pen JOIN produk pr ON pen.produk_id = pr.id
            LEFT JOIN users u ON pen.user_id = u.id
            LEFT JOIN users vu ON pen.void_by = vu.id
            WHERE {where} ORDER BY pen.created_at DESC
        """, params).fetchall()

        totals = db.execute(f"""
            SELECT COALESCE(SUM(pen.total), 0), COALESCE(SUM(pen.keuntungan), 0), COUNT(*)
            FROM penjualan pen WHERE {where} AND (pen.status IS NULL OR pen.status = 'active')
        """, params).fetchone()

        produk = db.execute("SELECT * FROM produk WHERE stok > 0 ORDER BY nama").fetchall()

        # Get pending approvals for bos/og
        approvals = []
        if request.state.user["role"] in ("bos", "og"):
            approvals = db.execute("""
                SELECT a.*, p.nama as produk_nama, u.nama as requested_by
                FROM approval a JOIN produk p ON a.produk_id = p.id
                JOIN users u ON a.karyawan_id = u.id
                WHERE a.status = 'pending' ORDER BY a.created_at DESC
            """).fetchall()

    return templates.TemplateResponse(request, "penjualan.html", {
            "request": request, "user": request.state.user, "penjualan": penjualan,
            "totals": totals, "periode": periode, "produk": produk, "approvals": approvals,
            "tgl_dari": tgl_dari, "tgl_sampai": tgl_sampai,
        })

@app.post("/penjualan/tambah")
@require_auth
def penjualan_tambah(request: Request, produk_id: int = Form(...), jumlah: int = Form(...),
                     harga_jual_used: float = Form(0),
                     nama_customer: str = Form("-"), alamat_customer: str = Form(""),
                     hp_customer: str = Form(""), email_customer: str = Form(""),
                     keterangan: str = Form(""), metode_bayar: str = Form("tunai"),
                     approval_id: int = Form(0), tempo_hari: int = Form(30)):
    with get_db() as db:
        produk = db.execute("SELECT * FROM produk WHERE id = ?", (produk_id,)).fetchone()
        if produk["stok"] < jumlah:
            return RedirectResponse("/penjualan?error=stok_kurang", status_code=303)

        harga = harga_jual_used if harga_jual_used > 0 else produk["harga_jual"]
        total = harga * jumlah
        keuntungan = (harga - produk["harga_modal"]) * jumlah
        no_invoice = generate_no_invoice(db)

        db.execute("""
            INSERT INTO penjualan (user_id, produk_id, jumlah, harga_satuan, harga_modal, total, keuntungan,
                                   keterangan, nama_customer, alamat_customer, hp_customer, email_customer, approval_id, metode_bayar, no_invoice)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (request.state.user["id"], produk_id, jumlah, harga, produk["harga_modal"], total, keuntungan,
              keterangan, nama_customer, alamat_customer, hp_customer, email_customer,
              approval_id if approval_id else None, metode_bayar, no_invoice))

        db.execute("UPDATE produk SET stok = stok - ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (jumlah, produk_id))

        db.execute("""
            INSERT INTO stok_mutasi (produk_id, tipe, jumlah, harga_satuan, user_id, keterangan)
            VALUES (?, 'keluar', ?, ?, ?, ?)
        """, (produk_id, jumlah, harga, request.state.user["id"], f"Penjualan - {keterangan}" if keterangan else "Penjualan"))

        penjualan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Notification
        db.execute("INSERT INTO notifikasi (tipe, pesan, link) VALUES ('penjualan', ?, ?)",
                   (f"Penjualan: {produk['nama']} x{jumlah} = Rp {total:,.0f}", f"/penjualan/nota/{penjualan_id}"))

        # Create hutang if method is hutang/tempo
        if metode_bayar in ["hutang", "tempo"]:
            jatuh_tempo = (datetime.now() + timedelta(days=tempo_hari)).strftime("%Y-%m-%d")
            db.execute("""
                INSERT INTO hutang (pelanggan_id, penjualan_id, jumlah, sudah_bayar, sisa, status, jatuh_tempo, keterangan)
                VALUES (NULL, ?, ?, 0, ?, 'belum', ?, ?)
            """, (penjualan_id, total, total, jatuh_tempo, f"Customer: {nama_customer}"))
            db.execute("INSERT INTO notifikasi (tipe, pesan, link) VALUES ('hutang', ?, ?)",
                       (f"Hutang baru: Rp {total:,.0f} - jatuh tempo {jatuh_tempo}", "/hutang"))

        nama = produk['nama']
        log_audit(db, request.state.user, "Tambah Penjualan", "penjualan", f"Produk: {nama} x{jumlah} = Rp {total:,.0f} ({metode_bayar})", request.client.host)

    check_low_stock()
    return RedirectResponse(f"/penjualan/nota/{penjualan_id}?print=1", status_code=303)

@app.post("/penjualan/batch")
@require_auth
def penjualan_batch(request: Request):
    """Process multi-item cart checkout"""
    import json, uuid
    user = request.state.user
    # Parse JSON body
    body = request._body if hasattr(request, '_body') else b''
    
    # This route is called from form with hidden fields
    # We'll handle it via form data instead
    return RedirectResponse("/penjualan", status_code=303)

@app.post("/penjualan/checkout")
@require_auth
async def penjualan_checkout(request: Request):
    """Process multi-item cart checkout via JSON"""
    import json, uuid
    user = request.state.user
    data = await request.json()
    
    items = data.get('items', [])
    metode_bayar = data.get('metode_bayar', 'tunai')
    nama_customer = data.get('nama_customer', '-')
    alamat_customer = data.get('alamat_customer', '')
    hp_customer = data.get('hp_customer', '')
    email_customer = data.get('email_customer', '')
    keterangan = data.get('keterangan', '')
    diskon = data.get('diskon', 0)
    tempo_hari = data.get('tempo_hari', 30)
    
    if not items:
        return {"success": False, "error": "Keranjang kosong"}
    
    batch_id = str(uuid.uuid4())[:8]
    nota_ids = []
    
    with get_db() as db:
        for item in items:
            produk_id = item['produk_id']
            jumlah = item['jumlah']
            harga = item['harga']  # Already the final price (bottom price if diskon applied)
            item_diskon = item.get('diskon_item', 0)  # Per-item diskon from frontend
            
            produk = db.execute("SELECT * FROM produk WHERE id = ?", (produk_id,)).fetchone()
            if not produk:
                continue
            if produk['stok'] < jumlah:
                return {"success": False, "error": f"Stok {produk['nama']} tidak cukup ({produk['stok']} tersisa)"}
            
            total = harga * jumlah
            keuntungan = (harga - produk['harga_modal']) * jumlah
            no_invoice = generate_no_invoice(db)
            
            db.execute("""
                INSERT INTO penjualan (user_id, produk_id, jumlah, harga_satuan, harga_modal, total, keuntungan,
                                       keterangan, nama_customer, alamat_customer, hp_customer, email_customer,
                                       metode_bayar, diskon, batch_id, status, no_invoice)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)
            """, (user['id'], produk_id, jumlah, harga, produk['harga_modal'], total, keuntungan,
                  keterangan, nama_customer, alamat_customer, hp_customer, email_customer,
                  metode_bayar, item_diskon, batch_id, no_invoice))
            
            pen_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            nota_ids.append(pen_id)
            
            # Update stock
            db.execute("UPDATE produk SET stok = stok - ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                       (jumlah, produk_id))
            
            # Log stock mutation
            db.execute("""INSERT INTO stok_mutasi (produk_id, tipe, jumlah, harga_satuan, user_id, keterangan)
                          VALUES (?, 'keluar', ?, ?, ?, ?)""",
                       (produk_id, jumlah, harga, user['id'], f"Penjualan batch {batch_id}"))
            
            # Create hutang if needed
            if metode_bayar in ['hutang', 'tempo']:
                jatuh_tempo = (datetime.now() + timedelta(days=tempo_hari)).strftime("%Y-%m-%d")
                db.execute("""INSERT INTO hutang (pelanggan_id, penjualan_id, jumlah, sudah_bayar, sisa, status, jatuh_tempo, keterangan)
                              VALUES (NULL, ?, ?, 0, ?, 'belum', ?, ?)""",
                           (pen_id, total, total, jatuh_tempo, f"Batch {batch_id}"))
        
        # Log audit
        total_all = sum(item['harga'] * item['jumlah'] for item in items)
        log_audit(db, user, "Penjualan Batch", "penjualan",
                  f"Batch {batch_id}: {len(items)} item = Rp {total_all:,.0f} ({metode_bayar})",
                  request.client.host)
    
    check_low_stock()
    # Return the first nota ID for redirect
    if nota_ids:
        return {"success": True, "nota_id": nota_ids[0], "batch_id": batch_id, "count": len(nota_ids)}
    return {"success": False, "error": "Gagal memproses"}

@app.get("/penjualan/nota/{id}", response_class=HTMLResponse)
@require_auth
def nota_page(request: Request, id: int, print: int = 0):
    with get_db() as db:
        pen = db.execute("""
            SELECT pen.*, pr.nama as produk_nama, pr.kode as produk_kode,
                   u.nama as user_nama
            FROM penjualan pen JOIN produk pr ON pen.produk_id = pr.id
            LEFT JOIN users u ON pen.user_id = u.id
            WHERE pen.id = ?
        """, (id,)).fetchone()
    return templates.TemplateResponse(request, "invoice.html", {
        "request": request, "user": request.state.user, "pen": pen, "auto_print": print
    })

@app.post("/penjualan/void/{id}")
@require_bos_or_og
def penjualan_void(request: Request, id: int, alasan: str = Form(...)):
    user = request.state.user
    with get_db() as db:
        pen = db.execute("SELECT * FROM penjualan WHERE id=? AND status='active'", (id,)).fetchone()
        if not pen:
            return RedirectResponse("/penjualan?error=not_found", status_code=303)
        # Mark as voided
        db.execute("""UPDATE penjualan SET status='voided', void_reason=?, void_by=?, void_at=CURRENT_TIMESTAMP WHERE id=?""",
                   (alasan, user["id"], id))
        # Restore stock
        db.execute("UPDATE produk SET stok = stok + ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                   (pen["jumlah"], pen["produk_id"]))
        # Log stock mutation
        db.execute("""INSERT INTO stok_mutasi (produk_id, tipe, jumlah, harga_satuan, user_id, keterangan)
                      VALUES (?, 'masuk', ?, ?, ?, ?)""",
                   (pen["produk_id"], pen["jumlah"], pen["harga_satuan"], user["id"], f"Void penjualan #{id}: {alasan}"))
        log_audit(db, user, "Void Penjualan", "penjualan", f"#{id}: {pen['produk_nama']} x{pen['jumlah']} - Alasan: {alasan}", request.client.host)
    return RedirectResponse("/penjualan?voided=ok", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: PELANGGAN
# ═══════════════════════════════════════════════════════════════════════
@app.get("/pelanggan", response_class=HTMLResponse)
@require_auth
def pelanggan_page(request: Request, q: str = ""):
    with get_db() as db:
        query = "SELECT * FROM pelanggan WHERE 1=1"
        params = []
        if q:
            query += " AND (nama LIKE ? OR telepon LIKE ?)"
            params.extend([f"%{q}%", f"%{q}%"])
        query += " ORDER BY nama ASC"
        pelanggan = db.execute(query, params).fetchall()
    return templates.TemplateResponse(request, "pelanggan.html", {
        "request": request, "user": request.state.user, "pelanggan": pelanggan, "q": q
    })

@app.post("/pelanggan/tambah")
@require_bos_or_og
def pelanggan_tambah(request: Request, nama: str = Form(...), alamat: str = Form(""),
                     telepon: str = Form(""), email: str = Form(""), catatan: str = Form("")):
    with get_db() as db:
        db.execute("INSERT INTO pelanggan (nama, alamat, telepon, email, catatan) VALUES (?, ?, ?, ?, ?)",
                   (nama, alamat, telepon, email, catatan))
    return RedirectResponse("/pelanggan", status_code=303)

@app.post("/pelanggan/edit/{id}")
@require_bos_or_og
def pelanggan_edit(request: Request, id: int, nama: str = Form(...), alamat: str = Form(""),
                   telepon: str = Form(""), email: str = Form(""), catatan: str = Form("")):
    with get_db() as db:
        db.execute("UPDATE pelanggan SET nama=?, alamat=?, telepon=?, email=?, catatan=? WHERE id=?",
                   (nama, alamat, telepon, email, catatan, id))
    return RedirectResponse("/pelanggan", status_code=303)

@app.get("/pelanggan/hapus/{id}")
@require_bos
def pelanggan_hapus(request: Request, id: int):
    with get_db() as db:
        db.execute("DELETE FROM pelanggan WHERE id=?", (id,))
    return RedirectResponse("/pelanggan", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: HUTANG/PIUTANG
# ═══════════════════════════════════════════════════════════════════════
@app.get("/hutang", response_class=HTMLResponse)
@require_auth
def hutang_page(request: Request, status: str = "semua"):
    with get_db() as db:
        query = """
            SELECT h.*, COALESCE(p.nama, pen.nama_customer, '-') as pelanggan_nama,
                   p.telepon as pelanggan_telp, pen.total as penjualan_total,
                   pen.nama_customer
            FROM hutang h
            LEFT JOIN pelanggan p ON h.pelanggan_id = p.id
            LEFT JOIN penjualan pen ON h.penjualan_id = pen.id
        """
        if status == "belum":
            query += " WHERE h.status = 'belum'"
        elif status == "sebagian":
            query += " WHERE h.status = 'sebagian'"
        elif status == "lunas":
            query += " WHERE h.status = 'lunas'"
        query += " ORDER BY h.jatuh_tempo ASC"
        hutang = db.execute(query).fetchall()

        total_belum = db.execute("SELECT COALESCE(SUM(sisa), 0) FROM hutang WHERE status != 'lunas'").fetchone()[0]

    return templates.TemplateResponse(request, "hutang.html", {
        "request": request, "user": request.state.user, "hutang": hutang,
        "status": status, "total_belum": total_belum
    })

@app.post("/hutang/bayar")
@require_auth
def hutang_bayar(request: Request, hutang_id: int = Form(...), jumlah: float = Form(...),
                 metode_bayar: str = Form("tunai"), tanggal: str = Form(""), keterangan: str = Form("")):
    try:
        with get_db() as db:
            hutang = db.execute("SELECT * FROM hutang WHERE id=?", (hutang_id,)).fetchone()
            if not hutang:
                return RedirectResponse("/hutang", status_code=303)
            sisa_baru = hutang["sisa"] - jumlah
            sudah_baru = hutang["sudah_bayar"] + jumlah
            if sisa_baru <= 0:
                status = "lunas"
                sisa_baru = 0
            elif sudah_baru > 0:
                status = "sebagian"
            else:
                status = "belum"

            db.execute("UPDATE hutang SET sudah_bayar=?, sisa=?, status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                       (sudah_baru, sisa_baru, status, hutang_id))
            db.execute("INSERT INTO pembayaran_hutang (hutang_id, jumlah, user_id, metode_bayar, keterangan) VALUES (?, ?, ?, ?, ?)",
                       (hutang_id, jumlah, request.state.user["id"], metode_bayar, keterangan))
            log_audit(db, request.state.user, "Bayar Hutang", "hutang",
                      f"Bayar Rp {jumlah:,.0f} ({metode_bayar}) - sisa Rp {sisa_baru:,.0f}",
                      request.client.host if request.client else "")

            # Get payment ID for invoice
            pembayaran_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Redirect to invoice page
        return RedirectResponse(f"/hutang/invoice/{pembayaran_id}", status_code=303)
    except Exception as e:
        return RedirectResponse("/hutang", status_code=303)

@app.get("/hutang/riwayat/{hutang_id}")
@require_auth
def hutang_riwayat(request: Request, hutang_id: int):
    with get_db() as db:
        riwayat = db.execute("""
            SELECT ph.*, u.nama as user_nama
            FROM pembayaran_hutang ph
            LEFT JOIN users u ON ph.user_id = u.id
            WHERE ph.hutang_id = ?
            ORDER BY ph.created_at DESC
        """, (hutang_id,)).fetchall()
        hutang = db.execute("""
            SELECT h.*, COALESCE(p.nama, pen.nama_customer, '-') as pelanggan_nama
            FROM hutang h
            LEFT JOIN pelanggan p ON h.pelanggan_id = p.id
            LEFT JOIN penjualan pen ON h.penjualan_id = pen.id
            WHERE h.id = ?
        """, (hutang_id,)).fetchone()
    return {"hutang": dict(hutang) if hutang else None,
            "riwayat": [dict(r) for r in riwayat]}

@app.get("/hutang/invoice/{pembayaran_id}", response_class=HTMLResponse)
@require_auth
def hutang_invoice(request: Request, pembayaran_id: int, print: int = 0):
    with get_db() as db:
        bayar = db.execute("""
            SELECT ph.*, h.jumlah as hutang_total, h.sudah_bayar, h.sisa as sisa_hutang,
                   h.status as hutang_status, h.jatuh_tempo,
                   COALESCE(p.nama, pen.nama_customer, '-') as pelanggan_nama,
                   p.telepon as pelanggan_telp, p.alamat as pelanggan_alamat,
                   u.nama as user_nama
            FROM pembayaran_hutang ph
            JOIN hutang h ON ph.hutang_id = h.id
            LEFT JOIN pelanggan p ON h.pelanggan_id = p.id
            LEFT JOIN penjualan pen ON h.penjualan_id = pen.id
            LEFT JOIN users u ON ph.user_id = u.id
            WHERE ph.id = ?
        """, (pembayaran_id,)).fetchone()

    if not bayar:
        return RedirectResponse("/hutang", status_code=303)

    is_lunas = bayar["hutang_status"] == "lunas"
    metode = "Transfer" if bayar["metode_bayar"] == "transfer" else "Tunai"

    return templates.TemplateResponse(request, "invoice_universal.html", {
        "request": request,
        "user": request.state.user,
        "title": "Bukti Pembayaran Hutang",
        "no_invoice": f"PAY-{pembayaran_id:06d}",
        "tanggal": bayar["created_at"] or datetime.now().strftime("%Y-%m-%d %H:%M"),
        "badge_class": "badge-green" if is_lunas else "badge-orange",
        "badge_text": "✅ LUNAS" if is_lunas else "⏳ CICILAN",
        "info_label": "Pelanggan",
        "info_value": bayar["pelanggan_nama"],
        "info_extra": f"Telp: {bayar['pelanggan_telp']}" if bayar["pelanggan_telp"] else "",
        "detail_items": [
            {"label": "Total Tagihan", "value": f"Rp {bayar['hutang_total']:,.0f}"},
            {"label": "Jumlah Bayar", "value": f"Rp {bayar['jumlah']:,.0f}"},
            {"label": "Metode Bayar", "value": f"{'🏦' if bayar['metode_bayar'] == 'transfer' else '💵'} {metode}"},
            {"label": "Sisa Tagihan", "value": f"Rp {bayar['sisa_hutang']:,.0f}"},
            {"label": "Jatuh Tempo", "value": bayar["jatuh_tempo"] or "-"},
            {"label": "Diproses oleh", "value": bayar["user_nama"] or "-"},
        ],
        "table_items": [],
        "table_columns": [],
        "total_rows": [
            {"label": "Dibayar Sekarang", "value": f"Rp {bayar['jumlah']:,.0f}", "color": "#16a34a"},
            {"label": "Total Sudah Dibayar", "value": f"Rp {bayar['sudah_bayar']:,.0f}", "color": "#2563eb"},
            {"label": "Sisa Tagihan", "value": f"Rp {bayar['sisa_hutang']:,.0f}", "color": "#dc2626" if bayar['sisa_hutang'] > 0 else "#16a34a"},
        ],
        "footer_note": "Simpan bukti pembayaran ini sebagai arahan",
        "printed_by": request.state.user.get("nama", "-"),
        "printed_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "auto_print": print,
    })
# ROUTES: STOK OPNAME (Full Version - Session-based)
# ═══════════════════════════════════════════════════════════════════════

@app.get("/opname", response_class=HTMLResponse)
@require_auth
def opname_list(request: Request, status: str = ""):
    """Halaman utama: daftar sesi opname"""
    with get_db() as db:
        user = request.state.user
        where = ["1=1"]
        params = []

        if status:
            where.append("s.status = ?")
            params.append(status)

        sessions = db.execute(f"""
            SELECT s.*, u.nama as user_nama,
                   a.nama as approver_nama,
                   (SELECT COUNT(*) FROM opname_item WHERE session_id = s.id) as total_item,
                   (SELECT COUNT(*) FROM opname_item WHERE session_id = s.id AND selisih != 0) as total_selisih
            FROM opname_session s
            LEFT JOIN users u ON s.user_id = u.id
            LEFT JOIN users a ON s.approved_by = a.id
            WHERE {" AND ".join(where)}
            ORDER BY s.created_at DESC
        """, params).fetchall()

        kategori_list = db.execute("SELECT * FROM kategori ORDER BY nama").fetchall()

        stats_draft = db.execute("SELECT COUNT(*) FROM opname_session WHERE status='draft'").fetchone()[0]
        stats_selesai = db.execute("SELECT COUNT(*) FROM opname_session WHERE status='selesai'").fetchone()[0]
        stats_disetujui = db.execute("SELECT COUNT(*) FROM opname_session WHERE status='disetujui'").fetchone()[0]
        stats_ditolak = db.execute("SELECT COUNT(*) FROM opname_session WHERE status='ditolak'").fetchone()[0]

    return templates.TemplateResponse(request, "opname.html", {
        "request": request, "user": user, "sessions": sessions,
        "kategori_list": kategori_list, "filter_status": status,
        "stats_draft": stats_draft, "stats_selesai": stats_selesai,
        "stats_disetujui": stats_disetujui, "stats_ditolak": stats_ditolak,
    })

@app.post("/opname/baru")
@require_auth
def opname_buat_baru(request: Request, catatan: str = Form(""), kategori_id: str = Form("")):
    """Buat sesi opname baru"""
    with get_db() as db:
        user = request.state.user
        cursor = db.execute(
            "INSERT INTO opname_session (tanggal, status, catatan, user_id) VALUES (DATE('now'), 'draft', ?, ?)",
            (catatan, user["id"])
        )
        session_id = cursor.lastrowid

        if kategori_id and kategori_id != "":
            produk_list = db.execute("SELECT * FROM produk WHERE kategori_id = ? ORDER BY nama", (kategori_id,)).fetchall()
        else:
            produk_list = db.execute("SELECT * FROM produk ORDER BY nama").fetchall()

        for p in produk_list:
            db.execute("""
                INSERT INTO opname_item (session_id, produk_id, stok_sistem, stok_fisik, selisih)
                VALUES (?, ?, ?, 0, 0)
            """, (session_id, p["id"], p["stok"]))

        log_audit(db, user, "Membuat Stok Opname Baru", "stok_opname",
                  f"Session #{session_id} | {len(produk_list)} item | {catatan or 'Tanpa catatan'}",
                  request.client.host if request.client else "")

    return RedirectResponse(f"/opname/{session_id}", status_code=303)

@app.get("/opname/{session_id}", response_class=HTMLResponse)
@require_auth
def opname_detail(request: Request, session_id: int, kategori: str = ""):
    """Halaman input stok fisik per item"""
    with get_db() as db:
        user = request.state.user
        session = db.execute("SELECT * FROM opname_session WHERE id = ?", (session_id,)).fetchone()
        if not session:
            return RedirectResponse("/opname", status_code=303)

        where = ["oi.session_id = ?"]
        params = [session_id]
        if kategori:
            where.append("k.nama = ?")
            params.append(kategori)

        items = db.execute(f"""
            SELECT oi.*, p.nama as produk_nama, p.kode as produk_kode, k.nama as kategori_nama
            FROM opname_item oi
            JOIN produk p ON oi.produk_id = p.id
            LEFT JOIN kategori k ON p.kategori_id = k.id
            WHERE {" AND ".join(where)}
            ORDER BY k.nama, p.nama
        """, params).fetchall()

        kategori_list = db.execute("""
            SELECT DISTINCT k.nama FROM kategori k
            JOIN produk p ON k.id = p.kategori_id
            ORDER BY k.nama
        """).fetchall()

        total_item = db.execute("SELECT COUNT(*) FROM opname_item WHERE session_id = ?", (session_id,)).fetchone()[0]
        sudah_diisi = db.execute("SELECT COUNT(*) FROM opname_item WHERE session_id = ? AND stok_fisik > 0", (session_id,)).fetchone()[0]
        total_selisih = db.execute("SELECT COUNT(*) FROM opname_item WHERE session_id = ? AND selisih != 0", (session_id,)).fetchone()[0]

    return templates.TemplateResponse(request, "opname_detail.html", {
        "request": request, "user": user, "session": session, "items": items,
        "kategori_list": kategori_list, "filter_kategori": kategori,
        "total_item": total_item, "sudah_diisi": sudah_diisi, "total_selisih": total_selisih,
    })

@app.post("/opname/item/simpan")
@require_auth
def opname_simpan_single(request: Request, session_id: int = Form(...), produk_id: int = Form(...),
                          stok_fisik: int = Form(0), keterangan: str = Form("")):
    """Simpan satu item opname (per-baris)"""
    with get_db() as db:
        user = request.state.user
        session = db.execute("SELECT * FROM opname_session WHERE id = ?", (session_id,)).fetchone()
        if not session or session["status"] not in ("draft",):
            return RedirectResponse("/opname", status_code=303)

        item = db.execute("SELECT * FROM opname_item WHERE session_id = ? AND produk_id = ?",
                          (session_id, produk_id)).fetchone()
        if not item:
            return RedirectResponse(f"/opname/{session_id}", status_code=303)

        selisih = stok_fisik - item["stok_sistem"]

        db.execute("""
            UPDATE opname_item SET stok_fisik = ?, selisih = ?, keterangan = ? WHERE id = ?
        """, (stok_fisik, selisih, keterangan, item["id"]))

        db.execute("UPDATE opname_session SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (session_id,))

        log_audit(db, user, "Input Stok Opname", "stok_opname",
                  f"Session #{session_id} | Produk #{produk_id} | Sistem: {item['stok_sistem']} → Fisik: {stok_fisik} | Selisih: {selisih}",
                  request.client.host if request.client else "")

    return RedirectResponse(f"/opname/{session_id}", status_code=303)

@app.post("/opname/{session_id}/selesai")
@require_auth
def opname_selesai(request: Request, session_id: int):
    """OG menandai opname selesai"""
    with get_db() as db:
        user = request.state.user
        session = db.execute("SELECT * FROM opname_session WHERE id = ?", (session_id,)).fetchone()
        if not session or session["status"] != "draft":
            return RedirectResponse("/opname", status_code=303)

        db.execute("UPDATE opname_session SET status = 'selesai', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (session_id,))

        total_selisih = db.execute("SELECT COUNT(*) FROM opname_item WHERE session_id = ? AND selisih != 0", (session_id,)).fetchone()[0]
        log_audit(db, user, "Stok Opname Selesai", "stok_opname",
                  f"Session #{session_id} | {total_selisih} item selisih | Menunggu approval Bos",
                  request.client.host if request.client else "")

    return RedirectResponse("/opname", status_code=303)

@app.post("/opname/{session_id}/approve")
@require_bos
def opname_approve(request: Request, session_id: int, aksi: str = Form("approve")):
    """Bos approve/reject opname → kalau approve, auto-koreksi stok"""
    with get_db() as db:
        user = request.state.user
        session = db.execute("SELECT * FROM opname_session WHERE id = ?", (session_id,)).fetchone()
        if not session or session["status"] != "selesai":
            return RedirectResponse("/opname", status_code=303)

        if aksi == "approve":
            items = db.execute("""
                SELECT oi.*, p.nama as produk_nama FROM opname_item oi
                JOIN produk p ON oi.produk_id = p.id WHERE oi.session_id = ?
            """, (session_id,)).fetchall()
            koreksi_count = 0
            for item in items:
                if item["selisih"] != 0:
                    db.execute("UPDATE produk SET stok = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                              (item["stok_fisik"], item["produk_id"]))
                    koreksi_count += 1

                    # Also log to stok_opname for history
                    db.execute("""
                        INSERT INTO stok_opname (produk_id, stok_sistem, stok_fisik, selisih, user_id, keterangan, session_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (item["produk_id"], item["stok_sistem"], item["stok_fisik"], item["selisih"],
                          session["user_id"], item["keterangan"], session_id))

            db.execute("""
                UPDATE opname_session SET status = 'disetujui', approved_by = ?, approved_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = ?
            """, (user["id"], session_id))

            log_audit(db, user, "Stok Opname Disetujui", "stok_opname",
                      f"Session #{session_id} | {koreksi_count} produk dikoreksi",
                      request.client.host if request.client else "")
        else:
            db.execute("""
                UPDATE opname_session SET status = 'ditolak', approved_by = ?, approved_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = ?
            """, (user["id"], session_id))

            log_audit(db, user, "Stok Opname Ditolak", "stok_opname",
                      f"Session #{session_id}",
                      request.client.host if request.client else "")

    return RedirectResponse("/opname", status_code=303)

@app.get("/opname/{session_id}/print", response_class=HTMLResponse)
@require_auth
def opname_print(request: Request, session_id: int, mode: str = "blank"):
    """Print view: blank (untuk cek manual) atau hasil (sudah diisi)"""
    with get_db() as db:
        session = db.execute("SELECT * FROM opname_session WHERE id = ?", (session_id,)).fetchone()
        if not session:
            return RedirectResponse("/opname", status_code=303)

        items = db.execute("""
            SELECT oi.*, p.nama as produk_nama, p.kode as produk_kode, k.nama as kategori_nama
            FROM opname_item oi
            JOIN produk p ON oi.produk_id = p.id
            LEFT JOIN kategori k ON p.kategori_id = k.id
            ORDER BY k.nama, p.nama
        """).fetchall()

        user_info = db.execute("SELECT nama FROM users WHERE id = ?", (session["user_id"],)).fetchone()

    return templates.TemplateResponse(request, "opname_print.html", {
        "request": request, "user": request.state.user, "session": session,
        "items": items, "mode": mode, "user_nama": user_info["nama"] if user_info else "-",
    })

@app.get("/opname/riwayat", response_class=HTMLResponse)
@require_auth
def opname_riwayat(request: Request, tgl_dari: str = "", tgl_sampai: str = ""):
    """Riwayat semua stok opname"""
    with get_db() as db:
        where = "1=1"
        params = []
        if tgl_dari:
            where += " AND DATE(o.created_at) >= ?"
            params.append(tgl_dari)
        if tgl_sampai:
            where += " AND DATE(o.created_at) <= ?"
            params.append(tgl_sampai)
        riwayat = db.execute(f"""
            SELECT o.*, p.nama as produk_nama, p.kode as produk_kode, u.nama as user_nama
            FROM stok_opname o JOIN produk p ON o.produk_id = p.id
            LEFT JOIN users u ON o.user_id = u.id
            WHERE {where} ORDER BY o.created_at DESC LIMIT 100
        """, params).fetchall()
    return templates.TemplateResponse(request, "opname_riwayat.html", {
        "request": request, "user": request.state.user, "riwayat": riwayat,
        "tgl_dari": tgl_dari, "tgl_sampai": tgl_sampai,
    })

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: LAPORAN
# ═══════════════════════════════════════════════════════════════════════
@app.get("/laporan", response_class=HTMLResponse)
@require_bos_or_og
def laporan_page(request: Request):
    with get_db() as db:
        bulan_ini = datetime.now().strftime("%Y-%m")
        harian = db.execute("""
            SELECT DATE(created_at) as tanggal, SUM(total) as total_penjualan,
                   SUM(keuntungan) as total_keuntungan, COUNT(*) as jumlah_transaksi
            FROM penjualan WHERE strftime('%Y-%m', created_at) = ?
            GROUP BY DATE(created_at) ORDER BY tanggal DESC
        """, (bulan_ini,)).fetchall()

        mingguan = []
        for i in range(4):
            start = (datetime.now() - timedelta(days=(i+1)*7)).strftime("%Y-%m-%d")
            end = (datetime.now() - timedelta(days=i*7)).strftime("%Y-%m-%d")
            row = db.execute("""
                SELECT COALESCE(SUM(total), 0), COALESCE(SUM(keuntungan), 0), COUNT(*)
                FROM penjualan WHERE DATE(created_at) >= ? AND DATE(created_at) < ?
            """, (start, end)).fetchone()
            mingguan.append({"minggu": f"{start} s/d {end}", "total": row[0], "keuntungan": row[1], "transaksi": row[2]})

        bulanan = []
        for i in range(6):
            d = datetime.now().replace(day=1) - timedelta(days=i*30)
            m = d.strftime("%Y-%m")
            row = db.execute("""
                SELECT COALESCE(SUM(total), 0), COALESCE(SUM(keuntungan), 0), COUNT(*)
                FROM penjualan WHERE strftime('%Y-%m', created_at) = ?
            """, (m,)).fetchone()
            bulanan.append({"bulan": m, "total": row[0], "keuntungan": row[1], "transaksi": row[2]})

    return templates.TemplateResponse(request, "laporan.html", {
        "request": request, "user": request.state.user, "harian": harian,
        "mingguan": mingguan, "bulanan": bulanan
    })

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: KATEGORI
# ═══════════════════════════════════════════════════════════════════════
@app.get("/kategori", response_class=HTMLResponse)
@require_auth
def kategori_page(request: Request):
    with get_db() as db:
        kategori = db.execute("""
            SELECT k.*, COUNT(p.id) as jumlah_produk FROM kategori k
            LEFT JOIN produk p ON p.kategori_id = k.id GROUP BY k.id ORDER BY k.nama
        """).fetchall()
    return templates.TemplateResponse(request, "kategori.html", {
        "request": request, "user": request.state.user, "kategori": kategori
    })

@app.post("/kategori/tambah")
@require_bos_or_og
def kategori_tambah(request: Request, nama: str = Form(...)):
    with get_db() as db:
        db.execute("INSERT OR IGNORE INTO kategori (nama) VALUES (?)", (nama,))
    return RedirectResponse("/kategori", status_code=303)

@app.get("/kategori/hapus/{id}")
@require_bos_or_og
def kategori_hapus(request: Request, id: int):
    with get_db() as db:
        db.execute("DELETE FROM kategori WHERE id=?", (id,))
    return RedirectResponse("/kategori", status_code=303)

@app.post("/kategori/edit/{id}")
@require_bos_or_og
def kategori_edit(request: Request, id: int, nama: str = Form(...)):
    with get_db() as db:
        db.execute("UPDATE kategori SET nama=? WHERE id=?", (nama, id))
    return RedirectResponse("/kategori", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: USER MANAGEMENT (BOS ONLY)
# ═══════════════════════════════════════════════════════════════════════
@app.get("/users", response_class=HTMLResponse)
@require_bos_or_og
def users_page(request: Request):
    with get_db() as db:
        users = db.execute("SELECT * FROM users ORDER BY role, nama").fetchall()
        # Get permissions for all users
        all_perms = {}
        for u in users:
            all_perms[u["id"]] = get_user_permissions(db, u["id"])
    return templates.TemplateResponse(request, "users.html", {
        "request": request, "user": request.state.user, "users": users,
        "all_perms": all_perms, "all_features": ALL_FEATURES, "role_defaults": ROLE_DEFAULTS,
    })

@app.post("/users/tambah")
@require_bos
def users_tambah(request: Request, username: str = Form(...), password: str = Form(...),
                nama: str = Form(...), role: str = Form("karyawan")):
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    with get_db() as db:
        db.execute("INSERT INTO users (username, password_hash, nama, role) VALUES (?, ?, ?, ?)",
                   (username, pw_hash, nama, role))
        new_user_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        init_user_permissions(db, new_user_id, role)
        log_audit(db, request.state.user, "Tambah User", "user", f"User: {username} ({role})", request.client.host)
    return RedirectResponse("/users", status_code=303)

@app.post("/users/edit/{id}")
@require_bos
def users_edit(request: Request, id: int, nama: str = Form(...), role: str = Form(...),
               password: str = Form(""), aktif: int = Form(1)):
    with get_db() as db:
        if password:
            pw_hash = hashlib.sha256(password.encode()).hexdigest()
            db.execute("UPDATE users SET nama=?, role=?, password_hash=?, aktif=? WHERE id=?",
                       (nama, role, pw_hash, aktif, id))
        else:
            db.execute("UPDATE users SET nama=?, role=?, aktif=? WHERE id=?", (nama, role, aktif, id))
        log_audit(db, request.state.user, "Edit User", "user", f"User ID {id}: {nama} ({role})", request.client.host)
    return RedirectResponse("/users", status_code=303)

@app.get("/users/hapus/{id}")
@require_bos
def users_hapus(request: Request, id: int):
    with get_db() as db:
        db.execute("DELETE FROM users WHERE id=?", (id,))
    return RedirectResponse("/users", status_code=303)

@app.get("/api/users/{user_id}/permissions")
@require_bos_or_og
def get_permissions_api(request: Request, user_id: int):
    """API: Get user permissions as JSON"""
    with get_db() as db:
        perms = get_user_permissions(db, user_id)
        if not perms:
            target = db.execute("SELECT role FROM users WHERE id = ?", (user_id,)).fetchone()
            if target:
                perms = ROLE_DEFAULTS.get(target["role"], {})
        return JSONResponse({"permissions": perms, "features": ALL_FEATURES})

@app.post("/users/{user_id}/permissions/save")
@require_bos_or_og
async def save_permissions(request: Request, user_id: int):
    """Save permissions from form submission"""
    form = await request.form()
    with get_db() as db:
        target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not target or target["role"] == "bos":
            return RedirectResponse("/users", status_code=303)
        db.execute("DELETE FROM user_permissions WHERE user_id = ?", (user_id,))
        for feature in ALL_FEATURES:
            enabled = 1 if form.get(feature) == "on" else 0
            db.execute("INSERT INTO user_permissions (user_id, feature, enabled) VALUES (?, ?, ?)",
                       (user_id, feature, enabled))
        enabled_list = [f for f in ALL_FEATURES if form.get(f) == "on"]
        log_audit(db, request.state.user, "Update Permission", "user",
                  f"User: {target['username']} | Enabled: {', '.join(enabled_list) or 'none'}",
                  request.client.host if request.client else "")
    return RedirectResponse("/users", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: BACKUP
# ═══════════════════════════════════════════════════════════════════════
@app.get("/backup", response_class=HTMLResponse)
@require_bos_or_og
def backup_page(request: Request):
    with get_db() as db:
        backups = db.execute("SELECT * FROM backups ORDER BY created_at DESC").fetchall()
    return templates.TemplateResponse(request, "backup.html", {
        "request": request, "user": request.state.user, "backups": backups
    })

@app.post("/backup/buat")
@require_bos_or_og
def backup_buat(request: Request):
    with get_db() as db:
        log_audit(db, request.state.user, "Backup Database", "sistem", "Backup manual dibuat", request.client.host)
    auto_backup()
    return RedirectResponse("/backup", status_code=303)

@app.get("/backup/download/{filename}")
@require_bos_or_og
def backup_download(request: Request, filename: str):
    filepath = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(filepath):
        return FileResponse(filepath, filename=filename)
    return RedirectResponse("/backup", status_code=303)

@app.get("/backup/restore/{filename}")
@require_bos_or_og
def backup_restore(request: Request, filename: str):
    filepath = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(filepath):
        # Backup current first
        auto_backup()
        shutil.copy2(filepath, DB_PATH)
    return RedirectResponse("/backup?restored=1", status_code=303)

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: AUDIT LOG
# ═══════════════════════════════════════════════════════════════════════
@app.get("/audit-log", response_class=HTMLResponse)
@require_bos_or_og
def audit_log_page(request: Request, q: str = "", kategori: str = "", tgl_dari: str = "", tgl_sampai: str = ""):
    with get_db() as db:
        query = "SELECT * FROM audit_log WHERE 1=1"
        params = []
        if q:
            query += " AND (username LIKE ? OR aksi LIKE ? OR detail LIKE ?)"
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
        if kategori:
            query += " AND kategori = ?"
            params.append(kategori)
        if tgl_dari:
            query += " AND DATE(created_at) >= ?"
            params.append(tgl_dari)
        if tgl_sampai:
            query += " AND DATE(created_at) <= ?"
            params.append(tgl_sampai)
        query += " ORDER BY created_at DESC LIMIT 500"
        logs = db.execute(query, params).fetchall()
        stats = {
            "total": len(logs),
            "login": len([l for l in logs if l['kategori'] == 'autentikasi']),
            "penjualan": len([l for l in logs if l['kategori'] == 'penjualan']),
            "stok": len([l for l in logs if l['kategori'] == 'stok']),
        }
        # Get unread count for base template
        notif_count = db.execute("SELECT COUNT(*) FROM notifikasi WHERE dibaca=0").fetchone()[0]
    return templates.TemplateResponse(request, "audit_log.html", {
        "request": request, "user": request.state.user, "logs": logs, "stats": stats,
        "q": q, "kategori": kategori, "tgl_dari": tgl_dari, "tgl_sampai": tgl_sampai,
        "notif_count": notif_count
    })

# ═══════════════════════════════════════════════════════════════════════
# ROUTES: EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════════════
@app.get("/export/{tipe}")
@require_bos_or_og
def export_excel(request: Request, tipe: str):
    wb = openpyxl.Workbook()
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def style_header(ws, cols):
        for ci, cn in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=cn)
            c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bdr
        ws.auto_filter.ref = ws.dimensions

    with get_db() as db:
        if tipe == "produk":
            ws = wb.active; ws.title = "Data Produk"
            style_header(ws, ["Kode", "Barcode", "Nama", "Kategori", "Harga Modal", "Harga Jual", "Stok", "Stok Min", "Satuan"])
            for ri, r in enumerate(db.execute("""
                SELECT p.kode, p.barcode, p.nama, COALESCE(k.nama,'-'), p.harga_modal, p.harga_jual, p.stok, p.stok_minimum, p.satuan
                FROM produk p LEFT JOIN kategori k ON p.kategori_id=k.id ORDER BY p.nama
            """).fetchall(), 2):
                for ci, v in enumerate(r, 1):
                    c = ws.cell(row=ri, column=ci, value=v); c.border = bdr
                    if ci in [5,6]: c.number_format = '#,##0'

        elif tipe == "penjualan":
            ws = wb.active; ws.title = "Penjualan"
            style_header(ws, ["Tanggal", "Kode", "Produk", "Pelanggan", "Kasir", "Jumlah", "Harga", "Total", "Modal", "Untung"])
            for ri, r in enumerate(db.execute("""
                SELECT pen.created_at, p.kode, p.nama, COALESCE(pl.nama,'-'), COALESCE(u.nama,'-'),
                       pen.jumlah, pen.harga_satuan, pen.total, pen.harga_modal, pen.keuntungan
                FROM penjualan pen JOIN produk p ON pen.produk_id=p.id
                LEFT JOIN pelanggan pl ON pen.pelanggan_id=pl.id
                LEFT JOIN users u ON pen.user_id=u.id ORDER BY pen.created_at DESC
            """).fetchall(), 2):
                for ci, v in enumerate(r, 1):
                    c = ws.cell(row=ri, column=ci, value=v); c.border = bdr
                    if ci in [7,8,9,10]: c.number_format = '#,##0'

        elif tipe == "keuntungan":
            ws = wb.active; ws.title = "Laporan Keuntungan"
            style_header(ws, ["Tanggal", "Total Penjualan", "Total Keuntungan", "Transaksi"])
            for ri, r in enumerate(db.execute("""
                SELECT DATE(created_at), SUM(total), SUM(keuntungan), COUNT(*)
                FROM penjualan WHERE strftime('%Y-%m',created_at)=? GROUP BY DATE(created_at) ORDER BY created_at DESC
            """, (datetime.now().strftime("%Y-%m"),)).fetchall(), 2):
                for ci, v in enumerate(r, 1):
                    c = ws.cell(row=ri, column=ci, value=v); c.border = bdr
                    if ci in [2,3]: c.number_format = '#,##0'

        elif tipe == "mutasi":
            ws = wb.active; ws.title = "Mutasi Stok"
            style_header(ws, ["Tanggal", "Kode", "Produk", "Tipe", "Jumlah", "Harga", "User", "Keterangan"])
            for ri, r in enumerate(db.execute("""
                SELECT m.created_at, p.kode, p.nama, m.tipe, m.jumlah, m.harga_satuan, COALESCE(u.nama,'-'), m.keterangan
                FROM stok_mutasi m JOIN produk p ON m.produk_id=p.id
                LEFT JOIN users u ON m.user_id=u.id ORDER BY m.created_at DESC
            """).fetchall(), 2):
                for ci, v in enumerate(r, 1):
                    c = ws.cell(row=ri, column=ci, value=v); c.border = bdr

        elif tipe == "hutang":
            ws = wb.active; ws.title = "Hutang Piutang"
            style_header(ws, ["Pelanggan", "Total", "Sudah Bayar", "Sisa", "Status", "Jatuh Tempo", "Keterangan"])
            for ri, r in enumerate(db.execute("""
                SELECT p.nama, h.jumlah, h.sudah_bayar, h.sisa, h.status, h.jatuh_tempo, h.keterangan
                FROM hutang h JOIN pelanggan p ON h.pelanggan_id=p.id ORDER BY h.status, h.jatuh_tempo
            """).fetchall(), 2):
                for ci, v in enumerate(r, 1):
                    c = ws.cell(row=ri, column=ci, value=v); c.border = bdr
                    if ci in [2,3,4]: c.number_format = '#,##0'

    filename = f"export_{tipe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return FileResponse(filepath, filename=filename,
                       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════════════════
# Run
# ═══════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # Auto backup on start
    auto_backup()
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
